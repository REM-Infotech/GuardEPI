from flask import *
from flask_login import login_required
from app import app
from app import db

import os
import inspect
from typing import Type
from datetime import datetime
from sqlalchemy import DateTime


import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from werkzeug.utils import secure_filename

from app.models import *
from app.Forms import *


tipo = db.Model

def get_models(tipo) -> Type[tipo]:
    
    models = {"Equipamentos": ProdutoEPI,
            "grade": GradeEPI}
    
    return models[tipo]

def get_variable_name(value):
    frame = inspect.currentframe().f_back
    for name, val in frame.f_locals.items():
        if name == value:
            return name
    return None

@app.route("/importacao_produto", methods=["POST"])
@login_required
def importacao_produto():

    try:
        form = IMPORTEPIForm()

        if form.validate_on_submit():

            importe_epi(form)
            return redirect(f'{request.referrer}')

        if form.errors:

            for erros in list(form.errors):
                message = form.errors[erros][0]
                flash(message, "success")

    except Exception as e:
        print(e)
        abort(500)

def importe_epi(form):

    doc =  form.arquivo.data
    docname = secure_filename(doc.filename)

    doc.save(os.path.join(app.config['Docs_Path'], f"{docname}"))
    doc_path = os.path.join(app.config['Docs_Path'], f"{docname}")

    
    update_counts(doc_path)

def update_counts(doc_path):
    
    model = get_models("Equipamentos")

    wb: Workbook = openpyxl.load_workbook(filename=doc_path)
    ws: Worksheet = wb.active

    for i in range(2, ws.max_row+1):
            
        id = ws.cell(row=i, column=1).value
        ca = ws.cell(row=i, column=2).value
        cod_ca = ws.cell(row=i, column=3).value
        nome_epi = ws.cell(row=i, column=4).value
        tipo_epi = ws.cell(row=i, column=5).value
        valor_unitario = ws.cell(row=i, column=6).value
        qtd_entregar = ws.cell(row=i, column=9).value
        periodicidade_item = ws.cell(row=i, column=10).value
        marca = ws.cell(row=i, column=11).value
        
        appends = [id, ca, cod_ca, nome_epi, tipo_epi, valor_unitario, qtd_entregar, periodicidade_item, marca]
        
        kwargs = {}
        
        for its in appends:
            for column in model.__table__.columns:
                
                
                default = ""
                if isinstance(column.type, DateTime):
                    default = datetime.now()
            
                var_name = get_variable_name(column.name)
                
                if var_name == column.name:
                    default = its
                
                    if kwargs.get(column.name, None) is None and default != "":
                        kwargs.setdefault(column.name, default)
                        break
                
        record = model.query.filter_by(nome_epi = kwargs.get("nome_epi")).first()
        
        if record is None:
            new_record = model(**kwargs)
            db.session.add(new_record)
            db.session.commit()

        if i == ws.max_row:
            flash("Equipamentos cadastrados com sucesso!", "success")

    db.session.commit()
