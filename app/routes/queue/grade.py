from flask import *
from flask_login import login_required
from app import app
from app import db

import os
import openpyxl
from sqlalchemy import func
from werkzeug.utils import secure_filename

from app.models import *
from app.Forms import *

databases = {
    "grade": [GradeEPI]
}

columns = {
    "grade": [GradeEPI.ca, GradeEPI.cod_ca, GradeEPI.nome_epi, GradeEPI.tipo_grade, GradeEPI.tipo_qtd, GradeEPI.qtd_estoque]        
}


@app.route("/importacao_grade", methods=["POST"])
@login_required
def importacao_grade():

    try:
        form = IMPORTEPIForm()

        if form.validate_on_submit():

            importe_epi(form)
            return redirect(f'{request.referrer}')

        if form.errors:

            for erros in list(form.errors):
                message = form.errors[erros][0]
                flash(message, "success")

    except Exception as e:
        abort(500)


def importe_epi(form):

    doc =  form.arquivo.data
    docname = secure_filename(doc.filename)

    doc.save(os.path.join(app.config['Docs_Path'], f"{docname}"))
    doc_path = os.path.join(app.config['Docs_Path'], f"{docname}")
    
    lista_itens = list(databases)
    
    for item in lista_itens:
        
        queryit = databases[item][0]
        queryobjs = columns[item]
        type_query = item
        update_counts(queryit, queryobjs, doc_path, type_query)
        
        set_estoque()
        
def update_counts(model, field, doc_path, type_query):
    

    wb = openpyxl.load_workbook(filename=doc_path)
    ws = wb.active

    for i in range(2, ws.max_row+1):

        ca = ws.cell(row=i, column=1).value
        cod_ca = ws.cell(row=i, column=2).value
        nome_epi = ws.cell(row=i, column=3).value
        tipo_grade =  ws.cell(row=i, column=4).value
        tipo_qtd = ws.cell(row=i, column=5).value
        qtd_estoque = ws.cell(row=i, column=6).value
        
        appends = [ca, cod_ca, nome_epi, tipo_grade, tipo_qtd, qtd_estoque]
        
        kwargs = {}
        
        for pos, its in enumerate(appends):
            
            fd = field[pos] 
            kwargs[fd.name] = its
                
        record = model.query.filter_by(nome_epi = kwargs.get("nome_epi")).first()
        
        if record is None or record.tipo_grade != tipo_grade:
            new_record = model(**kwargs)
            db.session.add(new_record)
            db.session.commit()
            
            
        

        if i == ws.max_row:
            flash("Equipamentos cadastrados com sucesso!", "success")

    db.session.commit()
    
def set_estoque():
    
    epis_agrupados = db.session.query(
    GradeEPI.nome_epi, GradeEPI.ca, GradeEPI.cod_ca, GradeEPI.tipo_qtd,
    func.sum(GradeEPI.qtd_estoque).label('quantidade_total')
    ).group_by(GradeEPI.nome_epi).all()

    # Atualizar ou inserir no banco de estoque
    for epi in epis_agrupados:
        estoque_item = EstoqueEPI.query.filter_by(nome_epi=epi.nome_epi).first()
        if estoque_item:
            estoque_item.quantidade_total = epi.quantidade_total
        else:
            novo_item = EstoqueEPI(nome_epi=epi.nome_epi, qtd_total_estoque=epi.quantidade_total, tipo_qtd=epi.tipo_qtd, ca=epi.ca)
            db.session.add(novo_item)

    db.session.commit()